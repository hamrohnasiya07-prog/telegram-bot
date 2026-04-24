import aiohttp, os
from telegram import *
from telegram.ext import *
from datetime import datetime, timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

TOKEN = "8547255151:AAEy3ZZOCFTNlsCd943vrQsFOKKMsH497d0"
API = "https://script.google.com/macros/s/AKfycbzll_uznQE4MYgfjHu4wc26rzlPsR9wPPwj4k761CiOWFKPLf4IcCYfmHveJh_Nxhl5/exec"

USERS = {}

# ===== API =====
async def api(p):
    async with aiohttp.ClientSession() as s:
        async with s.get(API, params=p) as r:
            return await r.json()

# ===== DATE LIST =====
def dates():
    return [(datetime.now()-timedelta(days=i)).strftime("%Y-%m-%d") for i in range(7)]

# ===== MESSAGE =====
def format_msg(d,title):
    return f"""
📊 {title}

👥 Mijoz: {d['mijoz']}
📞 Qo‘ng‘iroq: {d['q']}
💰 To‘lov: {d['t']}
🚫 Bog‘lanmadi: {d['bog']}
⚖️ Sud: {d['sud']}
⚠️ Muamoli: {d['mu']}
"""

# ===== EXCEL DESIGN =====
def make_excel(rows, filename):
    wb = Workbook()
    ws = wb.active

    headers = ["Hodim","Mijoz","Qo‘ng‘iroq","To‘lov","Bog‘lanmadi","Sud","Muamoli"]
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="00C0FF", fill_type="solid")

    for r in rows:
        ws.append([
            r["hodim"],
            r["mijoz"],
            r["q"],
            r["t"],
            r["bog"],
            r["sud"],
            r["mu"]
        ])

    wb.save(filename)

# ===== AUTO REPORT =====
async def auto_report(context):
    today = datetime.now().strftime("%Y-%m-%d")
    rows = await api({"type":"daily_all","date":today})

    if not rows:
        return

    total={"mijoz":0,"q":0,"t":0,"bog":0,"sud":0,"mu":0}

    for r in rows:
        total["mijoz"]+=r["mijoz"]
        total["q"]+=r["q"]
        total["t"]+=r["t"]
        total["bog"]+=r["bog"]
        total["sud"]+=r["sud"]
        total["mu"]+=r["mu"]

    text = format_msg(total,f"KUNLIK ({today})")

    for u in USERS:
        try:
            await context.bot.send_message(chat_id=u,text=text)
        except:
            pass

# ===== START =====
async def start(u,c):
    USERS[u.effective_chat.id]=True
    kb=[["📅 Hodim"],["📄 Barcha hodimlar"]]
    await u.message.reply_text("Tanlang:",reply_markup=ReplyKeyboardMarkup(kb,resize_keyboard=True))

# ===== HANDLE =====
async def handle(u,c):
    chat=u.effective_chat.id
    t=u.message.text
    state=USERS.get(chat,{})

    if t=="📅 Hodim":
        USERS[chat]={"step":"hodim"}
        kb=[[x] for x in ["Zufarbek","Layla","Nazokat","Bibizaxro"]]
        return await u.message.reply_text("Hodim:",reply_markup=ReplyKeyboardMarkup(kb,resize_keyboard=True))

    if state.get("step")=="hodim":
        USERS[chat]={"step":"date","hodim":t}
        kb=[[d] for d in dates()]
        return await u.message.reply_text("Sana:",reply_markup=ReplyKeyboardMarkup(kb,resize_keyboard=True))

    if state.get("step")=="date":
        d=await api({"type":"daily_hodim","date":t,"hodim":state["hodim"]})
        return await u.message.reply_text(format_msg(d,f"{state['hodim']} ({t})"))

    if t=="📄 Barcha hodimlar":
        USERS[chat]={"step":"all_date"}
        kb=[[d] for d in dates()]
        return await u.message.reply_text("Sana:",reply_markup=ReplyKeyboardMarkup(kb,resize_keyboard=True))

    if state.get("step")=="all_date":
        rows=await api({"type":"daily_all","date":t})

        file=f"{t}.xlsx"
        make_excel(rows,file)

        await u.message.reply_document(document=open(file,"rb"))
        os.remove(file)

# ===== APP =====
app=ApplicationBuilder().token(TOKEN).build()

app.add_handler(CommandHandler("start",start))
app.add_handler(MessageHandler(filters.TEXT,handle))

# ⏰ AUTO 18:10
app.job_queue.run_daily(auto_report, time=time(hour=18, minute=10))

print("BOT ISHLAYAPTI")
app.run_polling()
